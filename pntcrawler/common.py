import os
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime


def get_page_source(url, chrome_driver_path):
    """
    Selenium을 사용하여 url에 대응하는 html을 반환합니다.

    Args:
        url (str): html을 얻어올 url
        chrome_driver_path (str): 크롬 드라이버 실행파일 절대경로

    Return:
        <class 'str'>
    """

    # Selenium Headless(창 숨김 모드) 설정
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('lang=ko_KR')

    try:
        # Selenium을 이용하여 소모품 정보 페이지 HTML 얻어오기
        path = chrome_driver_path
        driver = webdriver.Chrome(path, options=chrome_options)
        driver.implicitly_wait(5)
        driver.get(url)
        pagesource = driver.page_source
        driver.close()

        return pagesource

    except Exception as e:
        print("Selenium 에러!")
        print(e)


def get_crawl_data(dept='', model='', ip='',
                   toner_k='', toner_c='', toner_m='', toner_y='',
                   drum_k='', drum_c='', drum_m='', drum_y='',
                   note=''):
    """
    크롤링 결과를 입력받아 프린터 소모품 정보가 담긴 Dictionary 객체를 반환합니다.

    Args:
        values (int 또는 str): 프린터 소모품 정보

    Return:
        <class 'dict'>
    """

    data = {
        'dept': dept, 'model': model, 'ip': ip,
        'toner_k': toner_k, 'toner_c': toner_c, 'toner_m': toner_m, 'toner_y': toner_y,
        'drum_k': drum_k, 'drum_c': drum_c, 'drum_m': drum_m, 'drum_y': drum_y,
        'note': note
    }

    return data


def get_error_data(error, dept, model, ip):
    """
    크롤링 에러 정보를 반환합니다.

    Args:
        error (Exception): Exception 객체
        dept (str): 부서 명
        model (str): 프린터 모델 명
        ip (str): 프린터 IP 주소

    Return:
        <class 'dict'>
    """

    if 'list index out of range' in str(error):
        _error = f'{str(error)}: html을 가져올 수 없는 상황(통신 불가 등)이거나 잘못된 크롤링 함수 사용'
    else:
        _error = str(error)

    print(f"{dept} 부서의 {model} 크롤링 실패... 원인은 다음과 같습니다.")
    print(_error)

    data = get_crawl_data(dept=dept, model=model, ip=ip, note=_error)
    return data


def create_xlsx(data, save_folder="crawldata",
                strftime_param="%Y-%m-%d_%H-%M-%S", save_file_prefix="CrawlingData",
                toner_warning=50, toner_alert=25,
                drum_warning=30, drum_alert=20,
                warning_color="FFFFDF00", alert_color="FFFFFF00"):
    """
    크롤링 최종 결과를 엑셀파일로 수합하여 저장합니다.

    Args:
        data (list): 크롤링 수합 정보가 담긴 list (필수 지정)
        save_folder: 크롤링 결과를 저장할 폴더 명 (선택)
        strftime_param: 엑셀 파일명에 시간을 명시할 때 strftime() 함수에 적용할 매개변수 (선택)
        save_file_prefix: 엑셀 파일명에 붙일 접두어 (선택)
        toner_warning (int): 토너 주의 잔량 (선택)
        toner_alert (int): 토너 경고 잔량 (선택)
        drum_warning (int): 드럼 주의 잔량 (선택)
        drum_alert (int): 드럼 경고 잔량 (선택)
        warning_color (str): 주의 색상 (선택)
        alert_color (str): 경고 색상 (선택)

    Return:
        None
    """
    wb = Workbook()  # 워크북
    ws = wb.active  # 워크시트

    # 틀 고정 및 컬럼 명 찍기
    ws.freeze_panes = "A2"
    ws.append([
        '부서명', '모델', 'IP',
        '토너(K)', '토너(C)', '토너(M)', '토너(Y)',
        '드럼(K)', '드럼(C)', '드럼(M)', '드럼(Y)',
        '비고'
    ])

    # title = ws.row_dimensions[1]
    # title.font = Font(bold=True)
    # title.alignment = Alignment(horizontal='center', vertical='center')

    try:
        # 폴더 생성
        if not(os.path.isdir(save_folder)):
            os.makedirs(os.path.join(save_folder))

        # 각 부서별 토너 및 드럼 잔량 찍기
        for item in data:
            print(item["dept"] + " 부서의 " + item["model"] + " 소모품 정보 저장 중...")
            ws.append([
                # IP 열에 하이퍼링크 지정
                item["dept"], item["model"], f'=HYPERLINK("http://{item["ip"]}")',
                item["toner_k"], item["toner_c"], item["toner_m"], item["toner_y"],
                item["drum_k"], item["drum_c"], item["drum_m"], item["drum_y"],
                item["note"]
            ])

        # 토너 잔량이 주의 및 경고인 경우 색칠하기
        fill_range = ws[f"d2:g{ws.max_row}"]  # D~G열: 토너 잔량 표시 범위
        for row in fill_range:
            for cell in row:
                if cell.value is '':  # 공백 셀은 건너뜀
                    continue
                elif isinstance(cell.value, int) and cell.value <= toner_alert:  # 경고 표시
                    cell.fill = PatternFill(
                        start_color=alert_color, end_color=alert_color, fill_type='solid')
                elif isinstance(cell.value, int) and toner_alert < cell.value <= toner_warning:  # 주의 표시
                    cell.fill = PatternFill(
                        start_color=warning_color, end_color=warning_color, fill_type='solid')
                elif isinstance(cell.value, str):  # 토너 잔량이 int로 뜨지 않는 경우('교환시기' 등): 경고 표시
                    cell.fill = PatternFill(
                        start_color=alert_color, end_color=alert_color, fill_type='solid')

        # 드럼 잔량이 주의 및 경고인 경우 색칠하기
        fill_range = ws[f"h2:k{ws.max_row}"]  # H~K열: 드럼 잔량 표시 범위
        for row in fill_range:
            for cell in row:
                # 공백 셀은 건너뜀
                if cell.value is '':
                    continue
                # 경고 표시(신도, OKI)
                elif isinstance(cell.value, int) and cell.value <= drum_alert:
                    cell.fill = PatternFill(
                        start_color=alert_color, end_color=alert_color, fill_type='solid')
                # 주의 표시(신도, OKI)
                elif isinstance(cell.value, int) and drum_alert < cell.value <= drum_warning:
                    cell.fill = PatternFill(
                        start_color=warning_color, end_color=warning_color, fill_type='solid')
                # 양호가 아닌 셀에 경고 색상 표시(제록스)
                elif isinstance(cell.value, str) and '양호' not in cell.value:
                    cell.fill = PatternFill(
                        start_color=alert_color, end_color=alert_color, fill_type='solid')

        # 파일명 지정
        date = datetime.today().strftime(strftime_param)
        file_name = f"{save_folder}/{save_file_prefix}_{date}.xlsx"
        print(f'{file_name} 파일에 저장 완료!')

        # 파일 저장
        wb.save(file_name)
        wb.close()
        print("저장 완료!")

    except OSError as e:
        print("폴더 생성 실패... 원인은 다음과 같습니다.")
        print(e)

    except Exception as e:
        print("엑셀 파일 생성 실패... 원인은 다음과 같습니다.")
        print(e)
